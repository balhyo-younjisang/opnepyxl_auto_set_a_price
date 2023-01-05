from openpyxl import load_workbook
from tkinter import *
from openpyxl.utils.cell import coordinate_from_string
import time
import sys

window = Tk()

window.title("Automatic Excel Macro")
window.geometry("640x400")
window.resizable(False, False)


def get_data():
    start_cell = str(startEntry.get())
    end_cell = str(endEntry.get())
    if start_cell == "" or end_cell == "":
        print("올바른 입력이 아닙니다.")
        sys.exit('Program terminated due to error')

    # get a product calculator excel file
    workbook_price = load_workbook("./상품가격비교계산기.xlsx", data_only=True)
    worksheet_price = workbook_price['위너 가능성 상품의 사본']

    try:
        get_cells = worksheet_price[start_cell:end_cell]
        for row in get_cells:
            for cell in row:
                # get a name,price,category of items
                pt = str(coordinate_from_string(cell.coordinate)[1])
                item_category = worksheet_price['A' + pt].value  # item category
                item_name = cell.value  # item name
                item_price = worksheet_price['L' + pt].value

                print("상품의 정보를 가져왔습니다 : " + item_category, item_name, item_price)
                enter_platform_data(item_category, item_name, item_price)

        workbook_price.save("./상품가격비교계산기.xlsx")
        workbook_price.close()
    except ValueError:
        print("올바른 입력이 아닙니다. 프로그램이 강제 종료됩니다.")
        sys.exit('Program terminated due to error')


def enter_product_data(product_name, open_market_data, smart_store_data, kakao_store_data, we_market_data):
    # singular product excel file
    workbook_singular = load_workbook("./단일상품.xlsx", data_only=True)
    worksheet_singular = workbook_singular['온라인상품']
    product_order = 0

    print(product_name)
    for i in range(2, worksheet_singular.max_row):
        if worksheet_singular['E' + str(i)].value == product_name:
            product_order = i
    if product_order == 0:
        print("같은 이름의 상품을 발견하지 못했습니다.")
        return

    auction = "옥션"
    g_market = "지마켓"
    eleven_st = "11번가"
    we_market = "위메프"

    account = worksheet_singular['C' + str(product_order)].value

    if auction in account:
        worksheet_singular['J' + str(product_order)].value = open_market_data
        print('J' + str(product_order) + "값 수정 :" + str(open_market_data))
    elif g_market in account:
        worksheet_singular['J' + str(product_order)].value = open_market_data
        print('J' + str(product_order) + "값 수정 :" + str(open_market_data))
    elif eleven_st in account:
        worksheet_singular['J' + str(product_order)].value = open_market_data
        print('J' + str(product_order) + "값 수정 :" + str(open_market_data))
    elif we_market in account:
        worksheet_singular['J' + str(product_order)].value = we_market_data
        print('J' + str(product_order) + "값 수정 : " + str(we_market_data))

    workbook_singular.save("./단일상품.xlsx")
    workbook_singular.close()

    # common product excel file
    workbook_common = load_workbook("./일반상품.xlsx", data_only=True)
    worksheet_common = workbook_common['온라인상품']

    smart_store = "스마트스토어"
    lotte_on = "롯데ON"
    kakao_store = "카카오톡 스토어"

    account = worksheet_common['C' + str(product_order)].value

    product_order = 0
    for i in range(2, worksheet_common.max_row):
        if worksheet_common['E' + str(i)].value == product_name:
            product_order = i
    if product_order == 0:
        print("같은 이름의 상품을 발견하지 못했습니다.")
        return

    if smart_store in account:
        worksheet_common['J' + str(product_order)].value = smart_store_data
        print('J' + str(product_order) + "값 수정 :" + str(smart_store_data))
    elif lotte_on in account:
        worksheet_common['J' + str(product_order)].value = open_market_data
        print('J' + str(product_order) + "값 수정 :" + str(open_market_data))
    elif kakao_store in account:
        worksheet_common['J' + str(product_order)].value = kakao_store_data
        print('J' + str(product_order) + "값 수정 :" + str(kakao_store_data))

    worksheet_common.save("./일반상품.xlsx")
    worksheet_common.close()


def enter_platform_data(category, name, price):
    # get a platform calculator excel file
    workbook_platform = load_workbook("./플랫폼+계산기.xlsx", data_only=True)
    worksheet_platform = workbook_platform['계산기']

    product_count = 0
    for i in range(3, worksheet_platform.max_row):
        if worksheet_platform['B' + str(i)].value == name:
            product_count = i
            break

    worksheet_platform['B' + str(product_count)].value = name
    worksheet_platform['C' + str(product_count)].value = price
    worksheet_platform['D' + str(product_count)].value = category

    open_market_data = worksheet_platform['K' + str(product_count)].value  # g-market, auction, 11st, lotte-on
    smart_store_data = worksheet_platform['E' + str(product_count)].value  # smart_store
    kakao_store_data = worksheet_platform['I' + str(product_count)].value  # kakao_store
    we_market_data = worksheet_platform['O' + str(product_count)].value  # we_market

    enter_product_data(name, open_market_data, smart_store_data, kakao_store_data, we_market_data)

    workbook_platform.save("./플랫폼+계산기.xlsx")
    workbook_platform.close()


def start():
    get_data()
    print("모든 자동화 과정이 끝났습니다.")
    time.sleep(3)
    print("프로그램이 곧 종료됩니다.")
    time.sleep(3)
    sys.exit("Application closed")


startLabel = Label(window, text="자동화가 시작될 상품명의 셀 주소를 입력해주세요. ex) C7")
startLabel.pack()
startEntry = Entry(window)
startEntry.pack()

endLabel = Label(window, text="자동화가 끝날 상품명의 셀 주소를 입력해주세요. ex) C10")
endLabel.pack()
endEntry = Entry(window)
endEntry.pack()

start_button = Button(window, text="자동화 시작", command=start)
start_button.pack()

window.mainloop()
